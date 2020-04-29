from time import gmtime,strftime
from pytz import timezone
from datetime import datetime
import time
import requests
from openpyxl import *
import pandas as pd
from lxml import html
from bs4 import BeautifulSoup

India = timezone('Asia/Kolkata')
ind_time = datetime.now(India)


try:
	session_requests = requests.session()
	login_url = "http://172.17.1.242/zabbix/index.php"
	res = session_requests.get(login_url)
except:
	print("Connection failed. Please Check VPN or Wifi")
		
def servers():
	#date="11/04/2018"
	#date1="11_04_2018"
	#time="5:00PM"
	date1 = ind_time.strftime("%d_%m_%Y")
	date = ind_time.strftime("%d/%m/%Y")
	hour = ind_time.strftime("%I")
	time = get_time(hour)
	print(date,date1,time)
	filename="Bluesr_New_Prod_Server_hourly_report_"+date1+".xlsx"
	
	server1("BSLWSPORTALPROD1",date,time,date1,filename)
	server2("BSLWSPORTALPROD2",date,time,date1,filename)
	server3("BSLWSPORTALPROD3",date,time,date1,filename)
	server4("BSLWSPDBPROD",date,time,date1,filename)
	server5("BSLWSHTTPPROD1",date,time,date1,filename)
	server6("BSLWSHTTPPROD2",date,time,date1,filename)
	server7("BSLWSLBPROD",date,time,date1,filename)
	server8("BSLWSDMGRPROD",date,time,date1,filename)
	server9("BSLWSTDSPROD",date,time,date1,filename)
	server10("BSLWSPORTALIHSDEV",date,time,date1,filename)
	server11("BSLWSIMSDATABASEUAT",date,time,date1,filename)
	server12("BSLWSSTADB",date,time,date1,filename)	
		
def server1(server_name,date,time,date1,filename):	
	sheet_index=0
	host_id="10266"
	index = [1,29,0,55,16,20,15,9,52,53,51,24,45,36,3,33,10,21,8,13,57,50,43,39,48,38,44,34,49,25,6,27,2,42,41,22,22,59,30,4,58,14,47]
	mydict={}

	get_data(host_id,mydict,index,date,time,filename,sheet_index,server_name)
def server2(server_name,date,time,date1,filename):	
	sheet_index=1
	host_id="10263"
	index = [1,29,0,55,15,19,14,8,52,53,51,23,45,36,3,33,9,20,7,12,57,50,43,-99,-99,-99,39,48,38,44,34,49,25,5,27,2,42,41,21,21,59,30,24,58,13,47]
	mydict={}

	get_data(host_id,mydict,index,date,time,filename,sheet_index,server_name)

def server3(server_name,date,time,date1,filename):	
	sheet_index=2
	host_id="10260"
	index = [1,29,0,55,15,19,14,8,52,53,51,23,45,36,3,33,9,20,7,12,57,50,43,-99,-99,-99,39,48,38,44,34,49,25,5,27,2,42,41,21,21,59,30,24,58,13,47]
	mydict={}

	get_data(host_id,mydict,index,date,time,filename,sheet_index,server_name)

def server4(server_name,date,time,date1,filename):	
	sheet_index=3
	host_id="10257"
	index = [1,29,0,55,16,20,15,9,52,53,51,24,45,36,3,33,10,21,8,13,57,50,43,-99,-99,-99,39,48,38,44,34,49,25,6,27,2,42,41,22,22,59,30,4,58,14,47]
	mydict={}

	get_data(host_id,mydict,index,date,time,filename,sheet_index,server_name)

def server5(server_name,date,time,date1,filename):	
	sheet_index=4
	host_id="10258"
	index = [1,29,0,55,16,20,15,9,52,53,51,24,45,36,3,33,10,21,8,13,57,50,43,-99,-99,-99,39,48,38,44,34,49,25,6,27,2,42,41,22,22,59,30,4,58,14,47]
	mydict={}

	get_data(host_id,mydict,index,date,time,filename,sheet_index,server_name)

def server6(server_name,date,time,date1,filename):	
	sheet_index=5
	host_id="10259"
	index = [1,29,0,55,16,20,15,9,52,53,51,24,45,36,3,33,10,21,8,13,57,50,43,-99,-99,-99,39,48,38,44,34,49,25,6,27,2,42,41,22,22,59,30,4,58,14,47]
	mydict={}

	get_data(host_id,mydict,index,date,time,filename,sheet_index,server_name)

def server7(server_name,date,time,date1,filename):	
	sheet_index=6
	host_id="10256"
	index = [1,29,0,55,16,20,15,9,52,53,51,24,45,36,3,33,10,21,8,13,57,50,43,-99,-99,-99,39,48,38,44,34,49,25,6,27,2,42,41,22,22,59,-99,-99,30,4,58,14,47]
	mydict={}

	get_data(host_id,mydict,index,date,time,filename,sheet_index,server_name)

def server8(server_name,date,time,date1,filename):	
	sheet_index=7
	host_id="10261"
	index = [1,29,0,55,16,20,15,9,52,53,51,24,45,36,3,33,10,21,8,13,57,50,43,-99,-99,-99,39,48,38,44,34,49,25,6,27,2,42,41,22,22,59,30,4,58,14,47]
	mydict={}

	get_data(host_id,mydict,index,date,time,filename,sheet_index,server_name)

def server9(server_name,date,time,date1,filename):	
	sheet_index=8
	host_id="10262"
	index = [1,29,0,55,16,20,15,9,52,53,51,24,45,36,3,33,10,21,8,13,57,50,43,-99,-99,-99,39,48,38,44,34,49,25,6,27,2,42,41,22,22,59,30,4,58,14,47]
	mydict={}

	get_data(host_id,mydict,index,date,time,filename,sheet_index,server_name)

def server10(server_name,date,time,date1,filename):	
	sheet_index=9
	host_id="10265"
	index = [3,33,0,63,19,24,17,11,60,61,59,18,51,43,5,40,12,25,10,15,67,58,49,66,34,65,50,46,57,27,9,30,4,48,47,26,26,69,35,55,68,16,53]
	mydict={}

	get_data(host_id,mydict,index,date,time,filename,sheet_index,server_name)

def server11(server_name,date,time,date1,filename):	
	sheet_index=10
	host_id="10264"
	index = [20,77,44,68,9,52,75,2,18,4,85,10,63,80,0,60,22,12,47,21,89,84,23,88,64,69,83,15,38,37,1,74,48,16,82,25,25,70,58,56,43,7,14]
	mydict={}

	get_data(host_id,mydict,index,date,time,filename,sheet_index,server_name)

def server12(server_name,date,time,date1,filename):	
	sheet_index=11
	host_id="10267"
	index = [1,15,0,26,9,10,8,3,24,4,23,12,20,18,27,22,6,11,11,29,16,2,28,7,21]
	mydict={}

	get_data(host_id,mydict,index,date,time,filename,sheet_index,server_name)

		
def get_data(host_id,mydict,index,date,time,filename,sheet_index,server_name):
	print("\nGetting data from:"+server_name)
	request_type=["CPU","Memory","Filesystems","Network+interfaces"]
	for type in request_type:
		res = session_requests.get("http://172.17.1.242/zabbix/latest.php?fullscreen=0&hostids%5B%5D="+host_id+"&application="+type+"&select=&filter_set=1")
		soup = BeautifulSoup(res.content, "html.parser")	
		make_dictionary(mydict,soup)
	
	"""
	try:
		file = open(server_name+"_param.txt", "w")
		for key in mydict:
			file.write(key+"   :"+mydict[key]+"\n")
			#print(key+"   :"+mydict[key]+"\n")
		file.close()
	except IOError:
		print("Please Close Param File")
		exit(0)
	"""
	#print(len(index))
	final_list = []
	final_list.insert(0,date)
	final_list.insert(1,time)
	for i in range(len(index)):
		if index[i] == -99 :
			final_list.insert(i+2,"No Data")
		else:			
			final_list.insert((i+2),mydict.values()[index[i]])
			
	#print(final_list)
	
	write_to_excel(final_list,filename,sheet_index,time)
	#read_excel(final_list,filename,sheet_index,time)
	
	
def make_dictionary(mydict,soup):
	count=0			#skip rows which are not useful
	count1=0		#to iterate over rows
	
	for td in soup.findAll("td", class_=""):		
		count=count+1
		if count>4:
			count1=count1+1			

			if count1==2 or count1==4:				
				if count1==2:
					mydict[td.text] = ""
					temp=td.text
				else:
					mydict[temp] = td.text
			if count1==7:
				count1=0
			
	#print(len(mydict))
		
def write_to_excel(final_list,filename,sheet_index,time):
	#print(len(final_list))
	#print(final_list)
	#try:
	wb = load_workbook(filename)
	sheets = wb.sheetnames
	ws = wb[sheets[sheet_index]]
	max_col = count_cols(ws)
	
	current_row = get_row(time)
	for row in ws.iter_rows(min_row=current_row, min_col=1, max_row=current_row, max_col=max_col):
		count=0
		for cell in row:
			#print("\nCount:"+str(count))
			try:
				str = final_list[count]
				if str[-1:] == "%":								
					cell.value = str
										
				else:	
					cell.value = str
				#print(final_list[count])
				
			except IndexError:
				print("Not found:"+str(count))
				exit(0)
				#continue;
			count=count+1
	wb.save(filename)
	#except IOError:
	#	print("Please Close Excel File")
	#	exit(0)
	
def count_cols (ws):
	for row in ws.iter_rows(min_row=3,max_row=3,min_col=1):
		count=1
		for cell in row:			
			if cell.value is not None:				
				count = count + 1
	return count
	
def get_row (time):
	if time == "10:00AM":
		return 4
	elif time == "11:00AM":
		return 5
	elif time == "12:00PM":
		return 6
	elif time == "01:00PM" :
		return 7
	elif time == "02:00PM" :
		return 8
	elif time == "03:00PM" :
		return 9
	elif time == "04:00PM" :
		return 10
	elif time == "05:00PM" :
		return 11
	elif time == "06:00PM" :
		return 12

def read_excel(final_list,filename,sheet_index,time):
	wb = load_workbook(filename)
	sheets = wb.sheetnames
	ws = wb[sheets[sheet_index]]
	max_col = count_cols(ws)
		
	current_row = get_row(time)
	for row in ws.iter_rows(min_row=current_row, min_col=1, max_row=current_row, max_col=max_col):
		count=0
		for cell in row:
			#print("\nCount:"+str(count))
			try:
				print(cell.value)
				#print(final_list[count])
				
			except IndexError:
				print("Not found:"+str(count))
				exit(0)
				#continue;
			count=count+1
		wb.save(filename)

def get_time(hour):
	if int(hour) == 10 or int(hour) == 11:
		return hour+":00AM"
	else:
		return hour+":00PM"
	
servers()	
	